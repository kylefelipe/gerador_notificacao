""" Criado para fazer a seleção da planilha a ser utilizada na montagem dos documentos """

from openpyxl import load_workbook

def  escolhe_folha(arquivo):
    """Realiza a escolha da aba dentro da planilha se a planilha tiver mais de uma aba dentro,
    é solicitado ao usuários que escolha o número da planilha na lista exibida """
    while True:
        try:
            aba_escolhida = 0
            abas = {}
            a=0

            if len(arquivo.worksheets) > 1:
                print('Essas são as {} abas encontradas' .format(len(arquivo.worksheets)))

            for i in arquivo.worksheets:
                b = (a, i.title)
                if len(arquivo.worksheets) > 1: print(b)
            #    print(b)
                abas.update([(a, i.title)])
                a += 1

            if len(arquivo.worksheets) > 1:
                aba_escolhida = (int(input("Digite número da aba escolhida: ")))
            abateste = abas[aba_escolhida]
            break
        except:
            print("Numero escolhido inválido. \n"
                  "Digite um dos números presentes na lista.")
    return abas[aba_escolhida]


def listar_documentos_contribuintes(tabela):

    """ Cria uma lista com cada CNPJ ou CPF dos contribuintes presentes na planilha removendo duplicatas"""

    todos_contribuintes = set()

    # Populando a lista
    for linha in tabela.iter_rows('A{}:A{}' .format(2 , tabela.max_row)):
        for celula in linha:
            todos_contribuintes.add(celula.value)
    contribuintes = list(todos_contribuintes)  # Transforma o set em lista
    return contribuintes


def listar_cnpjcpf_contribuintes(tabela):
    """ Cria um dicionário contendo o documento do contribuinte (cpf ou cnpj) como chave e o nome como siginificado"""

    contribuintes_documento_nome = {}
    for celula_1, celula_2 in tabela.iter_rows('A{}:B{}' .format(2, tabela.max_row)):
        contribuintes_documento_nome.update([('{}' .format(celula_1.value), '{}' .format(celula_2.value))])
    return contribuintes_documento_nome


def dados_empresa_tabela(cnpj, tabela_geral):
    """Retorna os itens da empresa, presentes na planilha, que devem constar na notificação,
       de acordo com o cnpj ou cpf"""

    dados_planilha = []  # Armazena os dados da empresa a serem inseridos na tabela

    # Definindo a area dos dados
    colunas_dados = tabela_geral.iter_rows('A{}:H{}' .format(2,tabela_geral.max_row))

    # Populando a lista aninhada (dados_planilha) com os dados da empresa(cnpj)
    for cnpj_emp, empresa, coluna_1, coluna_2, coluna_3, coluna_4, coluna_5, coluna_6 in colunas_dados:
        if cnpj_emp.value == cnpj:
            dados_planilha.append([coluna_1.value, coluna_2.value, coluna_3.value, coluna_4.value,
                                   coluna_5.value, coluna_6.value])
    return dados_planilha
